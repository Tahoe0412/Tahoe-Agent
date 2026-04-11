from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "docs" / "Tahoe_cost_budget_template.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="0F3D5E")
SECTION_FILL = PatternFill("solid", fgColor="DCEFF4")
SUBTLE_FILL = PatternFill("solid", fgColor="F4F8FA")
MONEY_FMT = '"$"#,##0.00'
PCT_FMT = "0.00%"
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E6EC"),
    right=Side(style="thin", color="D9E6EC"),
    top=Side(style="thin", color="D9E6EC"),
    bottom=Side(style="thin", color="D9E6EC"),
)


def style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_section(cell):
    cell.fill = SECTION_FILL
    cell.font = Font(bold=True, color="0F1720")
    cell.border = THIN_BORDER


def style_body(ws, start_row: int, end_row: int, end_col: int):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize(ws, widths: dict[str, int]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def freeze_and_filter(ws, freeze_cell: str, filter_range: str):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = filter_range


def add_cover_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "README"

    ws["A1"] = "Tahoe 成本与预算模板"
    ws["A1"].font = Font(size=16, bold=True, color="0F1720")
    ws["A3"] = "使用顺序"
    style_section(ws["A3"])

    steps = [
        "1. 先在 Parameters 填汇率、风险预留比例、默认固定成本。",
        "2. 在 Unit_Costs 维护各模块单次成本与平均版本次数。",
        "3. 在 Project_Estimator 按项目类型估算首版成本、试错成本、上架成本。",
        "4. 在 Monthly_Budget 输入当月各类型项目数量，查看预算。",
        "5. 在 Project_Ledger 逐项目登记实际发生额，月底复盘预算差异。",
    ]
    for idx, text in enumerate(steps, start=4):
        ws[f"A{idx}"] = text

    ws["A10"] = "专业口径说明"
    style_section(ws["A10"])
    notes = [
        "直接成本：文本、图片、视频、搜索 API 调用费。",
        "固定成本：服务器、数据库、SerpApi 月费、对象存储基础包等。",
        "研发费用：开发、设计、测试、运维工时。",
        "试错成本：改稿、重出图、重出视频、多轮 review 等返工成本。",
        "建议同时记录“首版成本”和“最终上架成本”，避免低估 AI 项目真实成本。",
    ]
    for idx, text in enumerate(notes, start=11):
        ws[f"A{idx}"] = text

    ws["A18"] = "项目类型定义"
    style_section(ws["A18"])
    definitions = [
        "TEXT_ONLY：纯文本交付，如主稿、标题包、平台改写、发布文案。",
        "TEXT_STORYBOARD：文本 + 分镜，但不直接出图 / 出视频。",
        "TEXT_IMAGE：文本 + 分镜 + 文生图。",
        "TEXT_IMAGE_VIDEO：文本 + 分镜 + 图片 + 视频，完整多模态交付。",
    ]
    for idx, text in enumerate(definitions, start=19):
        ws[f"A{idx}"] = text

    ws["A25"] = "文件说明"
    style_section(ws["A25"])
    ws["A26"] = "本模板已经预填 Tahoe 当前建议的质量优先模型分工，可直接修改。"

    autosize(ws, {"A": 110})


def add_parameters_sheet(wb: Workbook):
    ws = wb.create_sheet("Parameters")
    ws.append(["参数", "值", "说明"])
    style_header(ws[1])

    rows = [
        ("USD_CNY_RATE", 7.20, "美元兑人民币内部预算汇率"),
        ("RISK_BUFFER_RATE", 0.15, "风险预留比例，建议 10%-20%"),
        ("GPT54_INPUT_PER_1M", 2.50, "OpenAI GPT-5.4 官方输入单价 / 1M tokens"),
        ("GPT54_OUTPUT_PER_1M", 15.00, "OpenAI GPT-5.4 官方输出单价 / 1M tokens"),
        ("GPT54MINI_PROXY_INPUT_PER_1M", 0.25, "gpt-5.4-mini 预算口径暂按 OpenAI 官方 GPT-5 mini 输入价代理估算"),
        ("GPT54MINI_PROXY_OUTPUT_PER_1M", 2.00, "gpt-5.4-mini 预算口径暂按 OpenAI 官方 GPT-5 mini 输出价代理估算"),
        ("GEMINI31PRO_INPUT_PER_1M", 2.00, "Gemini 3 Pro Preview <=200K prompt 输入单价"),
        ("GEMINI31PRO_OUTPUT_PER_1M", 12.00, "Gemini 3 Pro Preview <=200K prompt 输出单价"),
        ("QWEN3MAX_INPUT_PER_1M", 1.20, "Qwen3-Max <=32K 输入单价"),
        ("QWEN3MAX_OUTPUT_PER_1M", 6.00, "Qwen3-Max <=32K 输出单价"),
        ("QWEN35PLUS_INPUT_PER_1M", 0.115, "Qwen3.5-Plus <=128K 输入单价"),
        ("QWEN35PLUS_OUTPUT_PER_1M", 0.688, "Qwen3.5-Plus <=128K 输出单价"),
        ("SERPER_PER_QUERY", 0.001, "Serper Starter: $50 / 50k credits"),
        ("SERPAPI_PER_QUERY", 0.025, "SerpApi Starter: $25 / 1k searches"),
        ("GEMINI_IMAGE_1K2K_PER_IMAGE", 0.134, "Gemini 3 Pro Image Preview 1K/2K 单张图成本"),
        ("GEMINI_IMAGE_4K_PER_IMAGE", 0.24, "Gemini 3 Pro Image Preview 4K 单张图成本"),
        ("VEO31_FAST_PER_SECOND", 0.15, "Veo 3.1 Fast 每秒视频成本"),
        ("VEO31_STANDARD_PER_SECOND", 0.40, "Veo 3.1 Standard 每秒视频成本"),
        ("MONTHLY_SERVER_COST_USD", 28.00, "预算假设：腾讯云轻量服务器月成本，未知时先按 $28 估"),
        ("MONTHLY_DB_STORAGE_COST_USD", 15.00, "预算假设：数据库 / 对象存储 / CDN 基础成本"),
        ("MONTHLY_SERPAPI_COST_USD", 25, "SerpApi 固定月费"),
        ("MONTHLY_OTHER_FIXED_COST_USD", 10.00, "其他固定技术成本预留"),
        ("MONTHLY_TOOL_SUBSCRIPTION_USD", 173.92, "按当前已知订阅折算：ChatGPT Plus $20 + Google Ultra $125 + See Dance 月摊 $28.92"),
        ("MONTHLY_RND_LABOR_USD", 0, "研发人工成本"),
        ("MONTHLY_QA_OPS_LABOR_USD", 0, "测试 / 运维 / 设计人工成本"),
    ]

    for row in rows:
        ws.append(row)

    style_body(ws, 2, ws.max_row, 3)
    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"].fill = SUBTLE_FILL
        if ws[f"A{row}"].value == "RISK_BUFFER_RATE":
            ws[f"B{row}"].number_format = PCT_FMT if ws[f"A{row}"].value == "RISK_BUFFER_RATE" else "0.00"
        elif "RATE" in str(ws[f"A{row}"].value):
            ws[f"B{row}"].number_format = "0.00"
        else:
            ws[f"B{row}"].number_format = MONEY_FMT

    autosize(ws, {"A": 34, "B": 18, "C": 60})


def add_unit_costs_sheet(wb: Workbook):
    ws = wb.create_sheet("Unit_Costs")
    ws.append([
        "模块编码",
        "生产深度",
        "成本分类",
        "模块名称",
        "推荐模型 / 服务",
        "单次成本低值(USD)",
        "单次成本高值(USD)",
        "平均版本次数",
        "默认说明",
        "预算低值(USD)",
        "预算高值(USD)",
    ])
    style_header(ws[1])

    rows = [
        ("SEARCH", "TEXT_ONLY", "直接成本", "搜索 / 资料整理", "Serper / SerpApi", 0.001, 0.025, 1.2, "按单次查询计价；低值按 Serper， 高值按 SerpApi"),
        ("TEXT_DRAFT", "TEXT_ONLY", "直接成本", "文本首稿", "Qwen3-Max / Gemini 3 Pro", 0.054, 0.098, 1.0, "约按 25K 输入 + 4K 输出估算"),
        ("TEXT_REVISION", "TEXT_ONLY", "试错成本", "文本改稿", "Qwen3.5-Plus / GPT-5.4", 0.004, 0.090, 2.0, "低值按轻量平台改写，高值按旗舰重审重写"),
        ("TEXT_REVIEW", "TEXT_ONLY", "试错成本", "质量 review", "GPT-5.4", 0.060, 0.138, 1.5, "约按 12K-25K 输入、2K-5K 输出估算"),
        ("PLATFORM_ADAPT", "TEXT_ONLY", "上架成本", "平台适配", "Qwen3.5-Plus / Qwen3-Max", 0.003, 0.026, 1.5, "小红书 / 抖音等派生改写"),
        ("SCRIPT_DRAFT", "TEXT_STORYBOARD", "直接成本", "脚本生成", "Gemini 3 Pro Preview", 0.196, 0.480, 1.5, "约按 50K-120K 输入、8K-20K 输出估算"),
        ("SCRIPT_REVISION", "TEXT_STORYBOARD", "试错成本", "脚本改稿", "Gemini 3 Pro Preview", 0.120, 0.284, 1.5, "结构性重写、补镜头、节奏调整"),
        ("STORYBOARD", "TEXT_STORYBOARD", "直接成本", "Storyboard 生成", "Gemini 3 Pro Preview", 0.308, 0.660, 1.5, "分镜与镜头描述生成"),
        ("SCENE_CLASSIFY", "TEXT_STORYBOARD", "直接成本", "Scene classification", "GPT-5.4 Mini(按 GPT-5 mini 代理)", 0.006, 0.016, 1.3, "透明说明：官方定价页未单列 gpt-5.4-mini，预算暂按 GPT-5 mini 代理"),
        ("ASSET_ANALYSIS", "TEXT_STORYBOARD", "直接成本", "Asset analysis", "GPT-5.4 Mini(按 GPT-5 mini 代理)", 0.009, 0.017, 1.3, "素材缺口、资产依赖分析"),
        ("REPORT_REVIEW", "TEXT_STORYBOARD", "试错成本", "报告 / review", "GPT-5.4", 0.110, 0.220, 1.5, "总结、诊断、复盘"),
        ("IMAGE_GEN", "TEXT_IMAGE", "直接成本", "文生图尝试", "Gemini 3 Pro Image Preview", 0.134, 0.240, 6.0, "低值按 1K/2K， 高值按 4K 单张图"),
        ("IMAGE_REWORK", "TEXT_IMAGE", "试错成本", "重绘 / 重出图", "Gemini 3 Pro Image Preview", 0.134, 0.240, 3.0, "二轮重绘、风格修正"),
        ("IMAGE_REVIEW", "TEXT_IMAGE", "上架成本", "图片最终 review", "GPT-5.4 Mini / GPT-5.4", 0.005, 0.040, 1.2, "选图、上架前检查"),
        ("VIDEO_GEN", "TEXT_IMAGE_VIDEO", "直接成本", "视频生成尝试", "Veo 3.1 Fast / Standard", 1.200, 3.200, 3.0, "按 8 秒视频估算：Fast $0.15/s，Standard $0.40/s"),
        ("VIDEO_REWORK", "TEXT_IMAGE_VIDEO", "试错成本", "视频返工", "Veo 3.1 Fast / Standard", 0.600, 2.400, 2.0, "按 4 秒 Fast 或 6 秒 Standard 返工估算"),
        ("FINAL_PACKAGE", "TEXT_IMAGE_VIDEO", "上架成本", "最终 packaging", "Qwen3.5-Plus / Qwen3-Max", 0.002, 0.021, 1.0, "标题、描述、封面包装"),
        ("FINAL_QA", "TEXT_IMAGE_VIDEO", "上架成本", "上架检查", "GPT-5.4 Mini / GPT-5.4", 0.006, 0.053, 1.0, "最终质检、平台检查"),
    ]

    for module_code, depth, cost_class, module_name, service, low_cost, high_cost, version_count, notes in rows:
        ws.append([
            module_code,
            depth,
            cost_class,
            module_name,
            service,
            low_cost,
            high_cost,
            version_count,
            notes,
            f"=F{ws.max_row + 1}*H{ws.max_row + 1}",
            f"=G{ws.max_row + 1}*H{ws.max_row + 1}",
        ])

    style_body(ws, 2, ws.max_row, 11)
    for col in ("F", "G", "J", "K"):
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = MONEY_FMT
    for row in range(2, ws.max_row + 1):
        ws[f"H{row}"].number_format = "0.0"

    freeze_and_filter(ws, "A2", f"A1:K{ws.max_row}")
    autosize(ws, {
        "A": 18,
        "B": 18,
        "C": 14,
        "D": 24,
        "E": 30,
        "F": 16,
        "G": 16,
        "H": 14,
        "I": 34,
        "J": 16,
        "K": 16,
    })


def add_project_estimator_sheet(wb: Workbook):
    ws = wb.create_sheet("Project_Estimator")
    ws["A1"] = "Tahoe 单项目成本测算"
    ws["A1"].font = Font(size=14, bold=True)

    headers = ["项目类型", "模块编码", "模块名称", "成本分类", "预算低值(USD)", "预算高值(USD)"]
    start_row = 3
    for offset, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=offset, value=header)
    style_header(ws[start_row])

    project_rows = [
        ("TEXT_ONLY", "SEARCH"),
        ("TEXT_ONLY", "TEXT_DRAFT"),
        ("TEXT_ONLY", "TEXT_REVISION"),
        ("TEXT_ONLY", "TEXT_REVIEW"),
        ("TEXT_ONLY", "PLATFORM_ADAPT"),
        ("TEXT_STORYBOARD", "SEARCH"),
        ("TEXT_STORYBOARD", "SCRIPT_DRAFT"),
        ("TEXT_STORYBOARD", "SCRIPT_REVISION"),
        ("TEXT_STORYBOARD", "STORYBOARD"),
        ("TEXT_STORYBOARD", "SCENE_CLASSIFY"),
        ("TEXT_STORYBOARD", "ASSET_ANALYSIS"),
        ("TEXT_STORYBOARD", "REPORT_REVIEW"),
        ("TEXT_IMAGE", "SEARCH"),
        ("TEXT_IMAGE", "SCRIPT_DRAFT"),
        ("TEXT_IMAGE", "SCRIPT_REVISION"),
        ("TEXT_IMAGE", "STORYBOARD"),
        ("TEXT_IMAGE", "SCENE_CLASSIFY"),
        ("TEXT_IMAGE", "ASSET_ANALYSIS"),
        ("TEXT_IMAGE", "REPORT_REVIEW"),
        ("TEXT_IMAGE", "IMAGE_GEN"),
        ("TEXT_IMAGE", "IMAGE_REWORK"),
        ("TEXT_IMAGE", "IMAGE_REVIEW"),
        ("TEXT_IMAGE_VIDEO", "SEARCH"),
        ("TEXT_IMAGE_VIDEO", "SCRIPT_DRAFT"),
        ("TEXT_IMAGE_VIDEO", "SCRIPT_REVISION"),
        ("TEXT_IMAGE_VIDEO", "STORYBOARD"),
        ("TEXT_IMAGE_VIDEO", "SCENE_CLASSIFY"),
        ("TEXT_IMAGE_VIDEO", "ASSET_ANALYSIS"),
        ("TEXT_IMAGE_VIDEO", "REPORT_REVIEW"),
        ("TEXT_IMAGE_VIDEO", "IMAGE_GEN"),
        ("TEXT_IMAGE_VIDEO", "IMAGE_REWORK"),
        ("TEXT_IMAGE_VIDEO", "VIDEO_GEN"),
        ("TEXT_IMAGE_VIDEO", "VIDEO_REWORK"),
        ("TEXT_IMAGE_VIDEO", "FINAL_PACKAGE"),
        ("TEXT_IMAGE_VIDEO", "FINAL_QA"),
    ]

    row = 4
    for project_type, module_code in project_rows:
        ws[f"A{row}"] = project_type
        ws[f"B{row}"] = module_code
        ws[f"C{row}"] = f'=IFERROR(VLOOKUP(B{row},Unit_Costs!$A:$K,4,FALSE),"")'
        ws[f"D{row}"] = f'=IFERROR(VLOOKUP(B{row},Unit_Costs!$A:$K,3,FALSE),"")'
        ws[f"E{row}"] = f'=IFERROR(VLOOKUP(B{row},Unit_Costs!$A:$K,10,FALSE),0)'
        ws[f"F{row}"] = f'=IFERROR(VLOOKUP(B{row},Unit_Costs!$A:$K,11,FALSE),0)'
        row += 1

    style_body(ws, 4, row - 1, 6)
    for col in ("E", "F"):
        for r in range(4, row):
            ws[f"{col}{r}"].number_format = MONEY_FMT

    summary_start = row + 2
    ws[f"A{summary_start}"] = "项目类型汇总"
    style_section(ws[f"A{summary_start}"])
    ws[f"A{summary_start+1}"] = "TEXT_ONLY"
    ws[f"A{summary_start+2}"] = "TEXT_STORYBOARD"
    ws[f"A{summary_start+3}"] = "TEXT_IMAGE"
    ws[f"A{summary_start+4}"] = "TEXT_IMAGE_VIDEO"
    ws[f"B{summary_start}"] = "预算低值(USD)"
    ws[f"C{summary_start}"] = "预算高值(USD)"
    style_header(ws[summary_start])
    for idx in range(1, 5):
        r = summary_start + idx
        ws[f"B{r}"] = f'=SUMIF($A$4:$A${row-1},A{r},$E$4:$E${row-1})'
        ws[f"C{r}"] = f'=SUMIF($A$4:$A${row-1},A{r},$F$4:$F${row-1})'
        ws[f"B{r}"].number_format = MONEY_FMT
        ws[f"C{r}"].number_format = MONEY_FMT

    freeze_and_filter(ws, "A4", f"A3:F{row-1}")
    autosize(ws, {"A": 20, "B": 18, "C": 28, "D": 16, "E": 16, "F": 16})


def add_monthly_budget_sheet(wb: Workbook):
    ws = wb.create_sheet("Monthly_Budget")
    ws["A1"] = "Tahoe 月预算"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = "项目预算"
    style_section(ws["A3"])
    ws.append(["项目类型", "月数量", "单项目预算低值(USD)", "单项目预算高值(USD)", "小计低值(USD)", "小计高值(USD)"])
    style_header(ws[4])

    project_types = ["TEXT_ONLY", "TEXT_STORYBOARD", "TEXT_IMAGE", "TEXT_IMAGE_VIDEO"]
    for idx, project_type in enumerate(project_types, start=5):
        ws[f"A{idx}"] = project_type
        ws[f"B{idx}"] = {
            "TEXT_ONLY": 40,
            "TEXT_STORYBOARD": 18,
            "TEXT_IMAGE": 10,
            "TEXT_IMAGE_VIDEO": 6,
        }[project_type]
        summary_row = {"TEXT_ONLY": 41, "TEXT_STORYBOARD": 42, "TEXT_IMAGE": 43, "TEXT_IMAGE_VIDEO": 44}[project_type]
        ws[f"C{idx}"] = f"=Project_Estimator!B{summary_row}"
        ws[f"D{idx}"] = f"=Project_Estimator!C{summary_row}"
        ws[f"E{idx}"] = f"=B{idx}*C{idx}"
        ws[f"F{idx}"] = f"=B{idx}*D{idx}"

    for col in ("C", "D", "E", "F"):
        for row in range(5, 9):
            ws[f"{col}{row}"].number_format = MONEY_FMT

    style_body(ws, 5, 8, 6)

    ws["A11"] = "固定成本预算"
    style_section(ws["A11"])
    ws.append(["固定成本项", "月预算(USD)", "说明"])
    style_header(ws[12])

    fixed_rows = [
        ("云服务器", "=Parameters!B3", "服务器月固定成本"),
        ("数据库 / 存储 / CDN", "=Parameters!B4", "数据库、存储、CDN"),
        ("SerpApi 固定月费", "=Parameters!B5", "固定 API 套餐"),
        ("其他固定技术成本", "=Parameters!B6", "其他基础设施"),
        ("工具订阅", "=Parameters!B7", "若要纳入平台预算可填写"),
        ("研发人工", "=Parameters!B8", "开发成本"),
        ("测试 / 运维 / 设计人工", "=Parameters!B9", "运营与交付支持"),
    ]
    for idx, row_data in enumerate(fixed_rows, start=13):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=idx, column=col_idx, value=value)
        ws[f"B{idx}"].number_format = MONEY_FMT

    style_body(ws, 13, 19, 3)

    ws["H3"] = "当前假设"
    style_section(ws["H3"])
    assumptions = [
        "项目量假设：TEXT_ONLY 40 / TEXT_STORYBOARD 18 / TEXT_IMAGE 10 / TEXT_IMAGE_VIDEO 6",
        "工具订阅已计入：ChatGPT Plus、Google AI Ultra(当前 $125/月)、See Dance 月摊",
        "Google AI Ultra 从 2026-05 起若涨到 $250/月，请把 Parameters 的工具订阅改为 298.92 USD",
        "服务器 / 数据库成本为预算占位值，后续可替换为真实账单",
    ]
    for idx, text in enumerate(assumptions, start=4):
        ws[f"H{idx}"] = text

    ws["A22"] = "预算汇总"
    style_section(ws["A22"])
    summary_labels = [
        ("项目直接成本低值", "=SUM(E5:E8)"),
        ("项目直接成本高值", "=SUM(F5:F8)"),
        ("固定成本合计", "=SUM(B13:B19)"),
        ("风险预留低值", "=($B$23+$B$25)*Parameters!B2"),
        ("风险预留高值", "=($B$24+$B$25)*Parameters!B2"),
        ("月总预算低值", "=B23+B25+B26"),
        ("月总预算高值", "=B24+B25+B27"),
        ("折人民币低值", "=B28*Parameters!B1"),
        ("折人民币高值", "=B29*Parameters!B1"),
    ]
    for idx, (label, formula) in enumerate(summary_labels, start=23):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = formula
        ws[f"B{idx}"].number_format = MONEY_FMT

    ws["B30"].number_format = MONEY_FMT
    ws["B31"].number_format = MONEY_FMT

    style_body(ws, 23, 31, 2)
    autosize(ws, {"A": 28, "B": 18, "C": 28, "D": 22, "E": 18, "F": 18, "H": 64})


def add_subscription_sheet(wb: Workbook):
    ws = wb.create_sheet("Subscriptions")
    ws.append(["订阅项", "币种", "当前金额", "折月金额(USD)", "计入口径", "备注"])
    style_header(ws[1])

    rows = [
        ("ChatGPT Plus", "USD", 20.00, 20.00, "工具订阅", "个人订阅；若平台统一使用可计入管理工具费"),
        ("Google AI Ultra", "USD", 125.00, 125.00, "工具订阅", "当前 2026-03/04 促销价；2026-05 起预计 $250/月"),
        ("See Dance", "CNY", 2499.00, 28.92, "工具订阅", "按年费 / 12 / 汇率 7.2 折算"),
        ("SerpApi", "USD", 25.00, 25.00, "固定技术成本", "平台固定 API 套餐"),
    ]
    for row in rows:
        ws.append(row)

    style_body(ws, 2, ws.max_row, 6)
    for row in range(2, ws.max_row + 1):
        ws[f"C{row}"].number_format = MONEY_FMT
        ws[f"D{row}"].number_format = MONEY_FMT

    ws["A8"] = "说明"
    style_section(ws["A8"])
    ws["A9"] = "如果你想只看平台 API / 交付成本，可在 Monthly_Budget 里把工具订阅行清零。"
    ws["A10"] = "如果你想看完整经营成本，则保留工具订阅、固定技术成本与人工。"

    autosize(ws, {"A": 20, "B": 10, "C": 14, "D": 16, "E": 16, "F": 44})


def add_project_ledger_sheet(wb: Workbook):
    ws = wb.create_sheet("Project_Ledger")
    ws.append([
        "项目ID",
        "项目名称",
        "业务线",
        "生产深度",
        "状态",
        "开始日期",
        "上架日期",
        "文本版本数",
        "图片尝试数",
        "视频尝试数",
        "搜索成本(USD)",
        "文本成本(USD)",
        "图片成本(USD)",
        "视频成本(USD)",
        "Review成本(USD)",
        "上架交付成本(USD)",
        "固定成本分摊(USD)",
        "总成本(USD)",
        "项目收入(USD)",
        "毛利(USD)",
        "备注",
    ])
    style_header(ws[1])

    for _ in range(15):
        ws.append(["", "", "", "", "", "", "", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "=SUM(K2:Q2)", 0, "=S2-R2", ""])

    for row in range(2, 17):
        ws[f"R{row}"] = f"=SUM(K{row}:Q{row})"
        ws[f"T{row}"] = f"=S{row}-R{row}"

    style_body(ws, 2, 16, 21)
    for col in ("K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"):
        for row in range(2, 17):
            ws[f"{col}{row}"].number_format = MONEY_FMT

    dv_line = DataValidation(type="list", formula1='"MARS_CITIZEN,MARKETING"', allow_blank=True)
    dv_depth = DataValidation(type="list", formula1='"TEXT_ONLY,TEXT_STORYBOARD,TEXT_IMAGE,TEXT_IMAGE_VIDEO"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"DRAFT,IN_PROGRESS,READY_TO_PUBLISH,PUBLISHED,ARCHIVED"', allow_blank=True)
    ws.add_data_validation(dv_line)
    ws.add_data_validation(dv_depth)
    ws.add_data_validation(dv_status)
    dv_line.add("C2:C200")
    dv_depth.add("D2:D200")
    dv_status.add("E2:E200")

    freeze_and_filter(ws, "A2", "A1:U200")
    autosize(ws, {
        "A": 14,
        "B": 26,
        "C": 16,
        "D": 18,
        "E": 18,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 18,
        "Q": 18,
        "R": 14,
        "S": 14,
        "T": 14,
        "U": 30,
    })


def add_pricing_sources_sheet(wb: Workbook):
    ws = wb.create_sheet("Pricing_Sources")
    ws.append(["日期", "类别", "项目", "价格摘记", "来源", "备注"])
    style_header(ws[1])

    rows = [
        ("2026-03-24", "OpenAI", "GPT-5.4", "Input $2.50 / 1M, Output $15.00 / 1M", "https://openai.com/api/pricing/", "官方 API 定价页"),
        ("2026-03-24", "OpenAI", "GPT-5 mini", "Input $0.25 / 1M, Output $2.00 / 1M", "https://openai.com/api/pricing/", "用于 gpt-5.4-mini 预算代理口径"),
        ("2026-03-24", "Gemini", "Gemini 3 Pro Preview", "Input $2 / 1M, Output $12 / 1M (<=200K prompts)", "https://ai.google.dev/pricing", "官方 Gemini Developer API 定价"),
        ("2026-03-24", "Gemini", "Gemini 3 Pro Image Preview", "1K/2K image about $0.134 each; 4K about $0.24 each", "https://ai.google.dev/gemini-api/docs/pricing?hl=fr", "官方定价页不同语言镜像，内容一致"),
        ("2026-03-24", "Gemini", "Veo 3.1", "Fast $0.15/s, Standard $0.40/s", "https://ai.google.dev/pricing", "官方 Gemini Developer API 定价"),
        ("2026-03-24", "Qwen", "qwen3-max", "Input $1.2 / 1M, Output $6 / 1M (<=32K)", "https://www.alibabacloud.com/help/en/model-studio/billing-for-model-studio", "官方 Model Studio 文档"),
        ("2026-03-24", "Qwen", "qwen3.5-plus", "Input $0.115 / 1M, Output $0.688 / 1M (<=128K)", "https://www.alibabacloud.com/help/en/model-studio/getting-started/models", "官方 Model Studio 文档"),
        ("2026-03-24", "Search", "Serper", "$50 / 50k credits = $0.001 per query", "https://serper.dev/", "官方站点 pricing"),
        ("2026-03-24", "Search", "SerpApi", "$25 / 1k searches = $0.025 per query", "https://serpapi.com/pricing", "官方 pricing"),
    ]

    for row in rows:
        ws.append(row)

    style_body(ws, 2, ws.max_row, 6)
    autosize(ws, {"A": 14, "B": 14, "C": 28, "D": 40, "E": 54, "F": 28})


def build_workbook():
    wb = Workbook()
    add_cover_sheet(wb)
    add_parameters_sheet(wb)
    add_unit_costs_sheet(wb)
    add_project_estimator_sheet(wb)
    add_monthly_budget_sheet(wb)
    add_project_ledger_sheet(wb)
    add_subscription_sheet(wb)
    add_pricing_sources_sheet(wb)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 22

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(f"Workbook written to {OUTPUT_PATH}")
